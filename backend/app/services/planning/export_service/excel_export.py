"""
Excel Export Module - Excel format export generation

Generates multi-sheet Excel workbooks with comprehensive planning data.

Architecture:
- Layer 2 (Service Layer): Format-specific export generation
- Uses openpyxl for Excel generation
"""

import io
import logging
from datetime import datetime
from typing import Any, Dict, Tuple
from uuid import UUID

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.context import RequestContext
from .base import BaseExportService

logger = logging.getLogger(__name__)


class ExcelExportService(BaseExportService):
    """
    Service for exporting planning data as Excel workbook.

    Generates multi-sheet workbook with:
    1. Summary - Executive summary and key metrics
    2. Waves - Wave plan details
    3. Resources - Resource allocations
    4. Timeline - Phase and milestone data
    5. Costs - Cost breakdown
    """

    def __init__(self, db: AsyncSession, context: RequestContext):
        """
        Initialize Excel export service.

        Args:
            db: Async SQLAlchemy database session
            context: Request context with tenant scoping
        """
        super().__init__(db, context)

    def export_excel(  # noqa: C901
        self, planning_data: Dict[str, Any], planning_flow_id: UUID
    ) -> Tuple[bytes, str, str]:
        """
        Export planning data as Excel workbook with multiple sheets.

        Args:
            planning_data: Complete planning data dictionary
            planning_flow_id: Planning flow UUID for filename

        Returns:
            Tuple of (excel_bytes, content_type, filename)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Define styles
            header_fill = PatternFill(
                start_color="1a73e8", end_color="1a73e8", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # === SHEET 1: Summary ===
            ws_summary = wb.create_sheet("Summary")

            wave_plan = planning_data.get("wave_plan", {})
            resource_allocation = planning_data.get("resource_allocation", {})
            cost_estimation = planning_data.get("cost_estimation", {})

            summary_data = [
                ["Migration Planning Summary", ""],
                ["", ""],
                ["Planning Flow ID", str(planning_data.get("planning_flow_id", "N/A"))],
                [
                    "Total Applications",
                    len(planning_data.get("selected_applications", [])),
                ],
                ["Total Waves", wave_plan.get("total_waves", 0)],
                ["Planning Status", planning_data.get("phase_status", "Unknown")],
                [
                    "Total Estimated Cost",
                    f"${cost_estimation.get('total_estimated_cost', 0):,.2f}",
                ],
                ["Export Date", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")],
            ]

            for row_idx, row_data in enumerate(summary_data, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:
                        cell.font = Font(bold=True, size=16, color="1a73e8")
                    elif col_idx == 1 and row_idx > 2:
                        cell.font = Font(bold=True)

            # Auto-adjust column widths
            ws_summary.column_dimensions["A"].width = 25
            ws_summary.column_dimensions["B"].width = 40

            # === SHEET 2: Waves ===
            ws_waves = wb.create_sheet("Waves")

            wave_headers = [
                "Wave Number",
                "Wave Name",
                "Applications Count",
                "Start Date",
                "End Date",
                "Dependencies",
            ]
            ws_waves.append(wave_headers)

            # Style headers
            for col_idx in range(1, len(wave_headers) + 1):
                cell = ws_waves.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Add wave data
            for wave in wave_plan.get("waves", []):
                wave_row = [
                    wave.get("wave_number", "N/A"),
                    wave.get("wave_name", "Unnamed"),
                    len(wave.get("applications", [])),
                    wave.get("start_date", "TBD"),
                    wave.get("end_date", "TBD"),
                    ", ".join(str(d) for d in wave.get("dependencies", [])),
                ]
                ws_waves.append(wave_row)

            # Auto-adjust columns
            for col_idx in range(1, len(wave_headers) + 1):
                ws_waves.column_dimensions[get_column_letter(col_idx)].width = 18

            # === SHEET 3: Resources ===
            ws_resources = wb.create_sheet("Resources")

            resource_headers = [
                "Wave",
                "Role",
                "Allocated Hours",
                "Hourly Rate",
                "Estimated Cost",
                "AI Suggested",
            ]
            ws_resources.append(resource_headers)

            # Style headers
            for col_idx in range(1, len(resource_headers) + 1):
                cell = ws_resources.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Add resource data
            for alloc in resource_allocation.get("allocations", []):
                wave_num = alloc.get("wave_number", "N/A")
                for resource in alloc.get("resources", []):
                    resource_row = [
                        f"Wave {wave_num}",
                        resource.get("role_name", "Unknown"),
                        resource.get("allocated_hours", 0),
                        f"${resource.get('hourly_rate', 0):.2f}",
                        f"${resource.get('estimated_cost', 0):,.2f}",
                        "Yes" if resource.get("is_ai_suggested", False) else "No",
                    ]
                    ws_resources.append(resource_row)

            # Auto-adjust columns
            for col_idx in range(1, len(resource_headers) + 1):
                ws_resources.column_dimensions[get_column_letter(col_idx)].width = 18

            # === SHEET 4: Timeline ===
            ws_timeline = wb.create_sheet("Timeline")

            timeline_headers = [
                "Phase Name",
                "Start Date",
                "End Date",
                "Status",
                "Progress (%)",
                "Dependencies",
            ]
            ws_timeline.append(timeline_headers)

            # Style headers
            for col_idx in range(1, len(timeline_headers) + 1):
                cell = ws_timeline.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Add timeline data
            for phase in planning_data.get("timeline", {}).get("phases", []):
                timeline_row = [
                    phase.get("name", "Unknown"),
                    phase.get("start_date", "TBD"),
                    phase.get("end_date", "TBD"),
                    phase.get("status", "Pending"),
                    f"{phase.get('progress', 0):.1f}%",
                    ", ".join(phase.get("dependencies", [])),
                ]
                ws_timeline.append(timeline_row)

            # Add milestones section
            milestones = planning_data.get("timeline", {}).get("milestones", [])
            if milestones:
                ws_timeline.append([])  # Empty row
                ws_timeline.append(["Milestones", "", "", "", "", ""])

                milestone_headers = [
                    "Milestone Name",
                    "Planned Date",
                    "Actual Date",
                    "Status",
                    "Phase",
                    "",
                ]
                ws_timeline.append(milestone_headers)

                for milestone in milestones:
                    milestone_row = [
                        milestone.get("name", "Unknown"),
                        milestone.get("planned_date", "TBD"),
                        milestone.get("actual_date", "N/A"),
                        milestone.get("status", "Pending"),
                        milestone.get("phase_id", "N/A"),
                        "",
                    ]
                    ws_timeline.append(milestone_row)

            # Auto-adjust columns
            for col_idx in range(1, len(timeline_headers) + 1):
                ws_timeline.column_dimensions[get_column_letter(col_idx)].width = 18

            # === SHEET 5: Costs ===
            ws_costs = wb.create_sheet("Costs")

            cost_headers = ["Category", "Amount"]
            ws_costs.append(cost_headers)

            # Style headers
            for col_idx in range(1, len(cost_headers) + 1):
                cell = ws_costs.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Add cost breakdown
            total_cost = cost_estimation.get("total_estimated_cost", 0)
            resource_cost = resource_allocation.get("total_cost_estimate", 0)
            contingency_pct = planning_data.get("planning_config", {}).get(
                "contingency_percentage", 15
            )

            cost_rows = [
                ["Total Estimated Cost", f"${total_cost:,.2f}"],
                ["Resource Costs", f"${resource_cost:,.2f}"],
                ["Contingency Buffer", f"{contingency_pct}%"],
                ["", ""],
            ]

            # Add per-wave costs if available
            for alloc in resource_allocation.get("allocations", []):
                wave_cost = sum(
                    r.get("estimated_cost", 0) for r in alloc.get("resources", [])
                )
                cost_rows.append(
                    [
                        f"Wave {alloc.get('wave_number', 'N/A')} Cost",
                        f"${wave_cost:,.2f}",
                    ]
                )

            for row_data in cost_rows:
                ws_costs.append(row_data)

            # Auto-adjust columns
            ws_costs.column_dimensions["A"].width = 30
            ws_costs.column_dimensions["B"].width = 20

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()

            # Generate filename
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"planning_export_{planning_flow_id}_{timestamp}.xlsx"

            logger.info(
                f"✅ Excel export generated: {len(excel_bytes)} bytes, 5 sheets"
            )

            return (
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename,
            )

        except Exception as e:
            logger.error(f"❌ Excel export failed: {e}", exc_info=True)
            raise
