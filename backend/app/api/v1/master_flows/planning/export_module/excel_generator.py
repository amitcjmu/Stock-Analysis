"""
Excel report generator for planning flow export (Issue #714).

Generates professional Excel reports using openpyxl.
"""

import io
from datetime import datetime
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _get_excel_styles():
    """Get common Excel styles."""
    header_fill = PatternFill(
        start_color="1e40af", end_color="1e40af", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    return header_fill, header_font, thin_border


def _create_summary_sheet(wb: Workbook, planning_flow: Any):
    """Create summary sheet for Excel report."""
    header_fill, header_font, thin_border = _get_excel_styles()

    ws = wb.active
    ws.title = "Summary"

    summary_data = [
        ["Planning Flow Report"],
        [],
        ["Property", "Value"],
        ["Planning Flow ID", str(planning_flow.planning_flow_id)],
        ["Master Flow ID", str(planning_flow.master_flow_id)],
        ["Current Phase", planning_flow.current_phase or "N/A"],
        ["Phase Status", planning_flow.phase_status or "N/A"],
        [
            "Created At",
            (
                planning_flow.created_at.strftime("%Y-%m-%d %H:%M:%S")
                if planning_flow.created_at
                else "N/A"
            ),
        ],
        [
            "Updated At",
            (
                planning_flow.updated_at.strftime("%Y-%m-%d %H:%M:%S")
                if planning_flow.updated_at
                else "N/A"
            ),
        ],
        ["Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")],
    ]

    for row_idx, row in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="1e40af")
            elif row_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = thin_border if row_idx >= 3 else None

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50


def _create_wave_sheet(wb: Workbook, planning_flow: Any):
    """Create wave plan sheet for Excel report."""
    header_fill, header_font, thin_border = _get_excel_styles()

    ws = wb.create_sheet("Wave Plan")
    wave_data = planning_flow.wave_plan_data or {}
    waves = wave_data.get("waves", []) if isinstance(wave_data, dict) else []

    wave_headers = [
        "Wave Name",
        "Asset Count",
        "Status",
        "Start Date",
        "End Date",
        "Duration (Days)",
    ]
    for col_idx, header in enumerate(wave_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, wave in enumerate(waves, 2):
        ws.cell(row=row_idx, column=1, value=wave.get("wave_name", "N/A")).border = (
            thin_border
        )
        ws.cell(row=row_idx, column=2, value=len(wave.get("assets", []))).border = (
            thin_border
        )
        ws.cell(row=row_idx, column=3, value=wave.get("status", "planned")).border = (
            thin_border
        )
        ws.cell(row=row_idx, column=4, value=wave.get("start_date", "TBD")).border = (
            thin_border
        )
        ws.cell(row=row_idx, column=5, value=wave.get("end_date", "TBD")).border = (
            thin_border
        )
        ws.cell(
            row=row_idx, column=6, value=wave.get("duration_days", "N/A")
        ).border = thin_border

    for col_idx in range(1, len(wave_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def _create_cost_sheet(wb: Workbook, planning_flow: Any):
    """Create cost estimation sheet for Excel report."""
    header_fill, header_font, thin_border = _get_excel_styles()

    ws = wb.create_sheet("Cost Estimation")
    cost_data = planning_flow.cost_estimation_data or {}

    cost_headers = ["Category", "Estimated Cost"]
    for col_idx, header in enumerate(cost_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row_idx = 2
    if isinstance(cost_data, dict):
        for key, value in cost_data.items():
            if key not in ("total", "currency"):
                ws.cell(
                    row=row_idx, column=1, value=key.replace("_", " ").title()
                ).border = thin_border
                cell = ws.cell(
                    row=row_idx,
                    column=2,
                    value=value if isinstance(value, (int, float)) else str(value),
                )
                cell.border = thin_border
                cell.number_format = (
                    "$#,##0.00" if isinstance(value, (int, float)) else "@"
                )
                row_idx += 1
        if cost_data.get("total"):
            cell_label = ws.cell(row=row_idx, column=1, value="TOTAL")
            cell_label.font = Font(bold=True)
            cell_label.border = thin_border
            cell_value = ws.cell(row=row_idx, column=2, value=cost_data["total"])
            cell_value.font = Font(bold=True)
            cell_value.border = thin_border
            cell_value.number_format = "$#,##0.00"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def _create_resource_sheet(wb: Workbook, planning_flow: Any):
    """Create resource allocation sheet for Excel report."""
    header_fill, header_font, thin_border = _get_excel_styles()

    ws = wb.create_sheet("Resource Allocation")
    resource_data = planning_flow.resource_allocation_data or {}

    res_headers = ["Resource Type", "Allocated", "Status"]
    for col_idx, header in enumerate(res_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row_idx = 2
    if isinstance(resource_data, dict):
        for key, value in resource_data.items():
            ws.cell(
                row=row_idx, column=1, value=key.replace("_", " ").title()
            ).border = thin_border
            ws.cell(
                row=row_idx,
                column=2,
                value=str(value) if not isinstance(value, dict) else "See details",
            ).border = thin_border
            ws.cell(row=row_idx, column=3, value="Allocated").border = thin_border
            row_idx += 1

    for col_idx in range(1, len(res_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25


def generate_excel_report(planning_flow: Any, metadata: Dict[str, Any]) -> io.BytesIO:
    """Generate Excel report for planning flow data (Issue #714)."""
    wb = Workbook()

    # Create all sheets
    _create_summary_sheet(wb, planning_flow)
    _create_wave_sheet(wb, planning_flow)
    _create_cost_sheet(wb, planning_flow)
    _create_resource_sheet(wb, planning_flow)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
